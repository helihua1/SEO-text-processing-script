"""
excel文档，取第一列作为关键词列表，关键词作为每次的user提示词。有一个txt文档里面是提示词作为system提示词。然后异步从deepseek api询问deepseek，限制最大异步量为20，当异步其中一个task完成以后，就传入一个下一个task，限制最大异步量为20。每次task得到的输出的内容，按照excel文档的顺序，放入同目录下的excel文档的对应行中。
功能：
从 Excel 第一列读取关键词，
使用 system.txt 作为 system prompt，
异步并发调用 DeepSeek API（最大并发 20），
按原顺序将结果写入 output.xlsx。
1. 每个 task 完成后即时写入 output.xlsx，不等待所有任务完成。
2. 支持 异步并发 20。
3. 每完成一个任务打印一次目前进度。
"""

import asyncio
import os
from openai import AsyncOpenAI
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

#deepseek api 的 key
api_key="sk-ff97f0da71e647eea56e78e84476ca88"

#提示词文件
system_prompt_file ="test.txt"

#输入文件
input_path ="input.xlsx"

#输出文件
file_path = "output.xlsx"



# ---------------------------------------------------------------
# 1. 读取 system prompt 和 keywords
# ---------------------------------------------------------------
def load_inputs():
    with open(system_prompt_file,  "r", encoding="utf-8") as f:
        system_prompt = f.read().strip()
        print("提示词",system_prompt)
    df = pd.read_excel(input_path, header=None)
    keywords = df.iloc[:, 0].astype(str).tolist()

    return system_prompt, keywords


# ---------------------------------------------------------------
# 2. 根据任务即时写入 Excel（不中断也能保存）
# ---------------------------------------------------------------
def write_one_row(idx, keyword, result):
    """
    idx: 行号（从0开始）
    keyword: 当前关键词
    result: deepseek 返回内容
    """


    # # 如果文件不存在：创建新的表头
    if not os.path.exists(file_path):
        wb = Workbook()
        ws = wb.active
        # ws.title = "result"
        # ws["A1"] = "Keyword"
        # ws["B1"] = "DeepSeek Output"
        wb.save(file_path)

    # 打开并写入第 idx+2 行（因为第一行为表头）
    wb = load_workbook(file_path)
    ws = wb.active

    row_num = idx + 1  # offset for header

    ws.cell(row=row_num, column=1, value=keyword)
    ws.cell(row=row_num, column=2, value=result)

    wb.save(file_path)


# ---------------------------------------------------------------
# 3. 单个任务：异步请求 DeepSeek
# ---------------------------------------------------------------
async def ask_deepseek(client, semaphore, system_prompt, keyword, idx, progress):
    async with semaphore:
        try:
            # 打印正在执行
            print(f"▶ 执行任务[{idx}]：{keyword}")

            resp = await client.chat.completions.create(
                model="deepseek-chat",
                messages=[
                    {"role": "system", "content": system_prompt},
                    {"role": "user", "content": keyword}
                ],
                stream=False
            )

            result = resp.choices[0].message.content
            print(keyword,'得到结果',result)
        except Exception as e:
            result = f"错误：{e}"

        # 每个任务完成后：立即写入 Excel
        write_one_row(idx, keyword, result)

        # 更新进度
        progress["done"] += 1
        print(f"✔ 完成 {progress['done']} / {progress['total']} : {keyword}")

        return idx, result


# ---------------------------------------------------------------
# 4. 主异步调度逻辑
# ---------------------------------------------------------------
async def async_main():
    print("程序启动...")

    # 4.1 加载输入
    system_prompt, keywords = load_inputs()
    total = len(keywords)

    # 4.2 DeepSeek 客户端
    client = AsyncOpenAI(
        api_key=api_key,
        base_url="https://api.deepseek.com"
    )

    # 4.3 控制最大并发 20
    semaphore = asyncio.Semaphore(20)

    # 进度记录
    progress = {"done": 0, "total": total}

    tasks = []
    for idx, kw in enumerate(keywords):
        task = asyncio.create_task(
            ask_deepseek(client, semaphore, system_prompt, kw, idx, progress)
        )
        tasks.append(task)

    # 等待全部任务结束
    await asyncio.gather(*tasks)

    print("全部任务完成！结果已写入 output.xlsx")


# ---------------------------------------------------------------
# 5. 启动程序
# ---------------------------------------------------------------
if __name__ == "__main__":
    asyncio.run(async_main())